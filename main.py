from openpyxl import load_workbook, Workbook

#   ABRINDO O ARQUIVO EXCEL
arquivo = load_workbook('alunos.xlsx')
planilha_alunos = arquivo['Alunos']

alunos_aprovados = []
alunos_reprovados = []

soma_notas = 0
for aluno in planilha_alunos.iter_rows(min_row=2, values_only=True):
    nome, curso, idade, nota_final, data_matricula = aluno
    
    aluno_dicionario = {
        'nome': nome,
        'curso': curso,
        'idade': idade,
        'nota_final': nota_final,
        'data_matricula': data_matricula
    }
    if nota_final >= 7:
        alunos_aprovados.append(aluno_dicionario)
    else:
        alunos_reprovados.append(aluno_dicionario)
soma_notas += nota_final
media_notas = soma_notas / (len(alunos_aprovados) + len(alunos_reprovados))

arquivo_aprovados = Workbook()
planilha_aprovados = arquivo_aprovados.active
planilha_aprovados.title = 'Aprovados'


arquivo_reprovados = Workbook()
planilha_reprovados = arquivo_reprovados.active
planilha_reprovados.title = 'Reprovados'



planilha_aprovados.append(['Nome', 'Curso', 'Idade', 'Nota Final', 'Data de Matrícula'])
for aluno in alunos_aprovados:
    planilha_aprovados.append(list(aluno.values()))

planilha_reprovados.append(['Nome', 'Curso', 'Idade', 'Nota Final', 'Data de Matrícula'])
for aluno in alunos_reprovados:
    planilha_reprovados.append(list(aluno.values()))

print(f'Alunos aprovados: {len(alunos_aprovados)}')
print(f'Alunos reprovados: {len(alunos_reprovados)}')
print(f'Média das notas: {media_notas:.2f}')


arquivo_aprovados.save('alunos_aprovados.xlsx')
arquivo_reprovados.save('alunos_reprovados.xlsx')